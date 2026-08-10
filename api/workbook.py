"""The Excel export: a LIVE workbook, not a printout of numbers.

Every figure past the input block is a formula, so an analyst can change an assumption and
watch it flow through the 52-month model. That carries one real risk — the workbook and the
Python engine computing different answers — and the whole design here is organized around
removing it.

How the two are kept locked together:

  * The formulas reproduce `engine/proforma.full_cashflow` line for line. Where the engine
    does `balance += draw + interest`, the sheet does `=prev + draw + interest`. There is no
    second model here, only the same model expressed in cells.
  * The vocabulary is deliberately boring: `+ - * / ^`, SUM, MIN, MAX, IF, IRR, and cell
    references. Nothing volatile, no array formulas, no functions whose edge cases differ
    between Excel, LibreOffice, and Sheets. An ugly sheet that agrees everywhere beats a
    clever one that does not.
  * `tests/test_xlsx_export.py` recomputes the generated workbook with an independent
    formula engine and asserts the Summary tab equals the Python engine to the cent.

Two things are NOT live, and both are labeled in the sheet rather than hidden:

  * **Residual land value.** The engine solves it with Brent's method — the land value at
    which levered IRR equals the hurdle. In a sheet that is circular (land -> costs -> loan
    -> equity -> IRR -> land). Rather than rely on iterative calculation, which converges
    differently across applications and can silently fail, the land cell is SEEDED with the
    solved value and the sheet computes IRR from it. The IRR cell reading exactly the hurdle
    is the in-sheet proof that the seed is right; edit an assumption and it drifts off,
    which is the signal to re-run Goal Seek (recipe is written next to the cell).
  * **The timeline and envelope.** Month count and program geometry fix the shape of the
    grid and are resolved upstream by `fit_program`, so changing them needs a re-export.
    They are shown in a separate block marked structural.

Sources & Uses reproduces `serializers.underwrite_response` EXACTLY, because matching the
tool is the trust property — the workbook must not quietly disagree with the panel it came
from. A check cell shows sources minus uses, which reads $0; when the §6.4 development-
equity scoping was fixed, this sheet inherited it by construction rather than by being
ported, which is what that coupling was for.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api import vocabulary as vocab
from engine.prototypes import (
    EXIT_CAP_ADJUSTMENT,
    MIN_EXIT_CAP_RATE,
    RENT_PREMIUM_FACTOR,
    hard_cost_psf,
)

# --- presentation ----------------------------------------------------------
MONEY = '"$"#,##0'
MONEY_2 = '"$"#,##0.00'
PCT = "0.00%"
NUM = "#,##0"
NUM_2 = "#,##0.00"
MULT = '0.00"×"'

INK = "FF1A1D1C"
ACCENT = "FF0E7C7B"
HEAD_FILL = PatternFill("solid", fgColor="FFEAF2F1")
INPUT_FILL = PatternFill("solid", fgColor="FFF5FAF9")
FIXED_FILL = PatternFill("solid", fgColor="FFF4F2ED")
WARN_FILL = PatternFill("solid", fgColor="FFFBF0E4")
THIN = Side(style="thin", color="FFD8D5CE")


def _title(ws, row: int, text: str, width: int = 8) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color=INK)
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = HEAD_FILL
        ws.cell(row=row, column=col).border = Border(bottom=THIN)
    return row + 1


def _note(ws, row: int, text: str, warn: bool = False) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(italic=True, size=9, color="FF8A5A21" if warn else "FF5C5952")
    cell.alignment = Alignment(wrap_text=False)
    return row + 1


class Sheet:
    """Writes labeled rows and remembers where each value landed.

    Formulas elsewhere reference these by address, so the layout can change without every
    formula in the workbook having to be rewritten by hand.
    """

    def __init__(self, ws, name: str):
        self.ws = ws
        self.name = name
        self.at: dict[str, str] = {}

    def label_value(
        self, row: int, key: str, label: str, value, fmt: str, fill=None, note: str = ""
    ) -> int:
        self.ws.cell(row=row, column=1, value=label).font = Font(size=10, color=INK)
        cell = self.ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.font = Font(size=10, bold=True, color=INK)
        if fill:
            cell.fill = fill
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if note:
            self.ws.cell(row=row, column=3, value=note).font = Font(size=9, color="FF8A8781")
        self.at[key] = f"'{self.name}'!$B${row}"
        return row + 1


def build(result: dict) -> Workbook:
    """Build the workbook for one `underwrite()` result."""
    program = result["program"]
    market = result["market"]
    parcel = result["parcel"]
    record = result["record"]
    assumptions = result["assumptions"]
    cf = result["cashflow"]
    outputs = result["outputs"]
    land_solved = result["full_rlv"]

    proto = program.prototype_id
    ctype = (
        program.construction_type.value
        if hasattr(program.construction_type, "value")
        else str(program.construction_type)
    )

    # Timeline, as the engine resolved it. `construction_months` is overridden to 30 for
    # concrete regardless of the assumption, so the EFFECTIVE value is what the grid uses.
    predev = int(assumptions.timeline["predevelopment_months"])
    stab = int(cf.phase_bounds["stabilization"])
    sale = int(cf.phase_bounds["sale"])
    construction = int(cf.phase_bounds["construction_end"]) - predev
    leaseup = stab - predev - construction
    months = sale + 1

    wb = Workbook()

    # =======================================================================
    # Assumptions
    # =======================================================================
    ws = wb.active
    ws.title = "Assumptions"
    a = Sheet(ws, "Assumptions")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 54

    r = 1
    ws.cell(row=r, column=1, value="Residual — model inputs").font = Font(bold=True, size=14)
    r += 1
    ws.cell(
        row=r, column=1,
        value=f"{record.get('address') or parcel.ssl} · {vocab.PROTOTYPE_LABELS.get(proto, proto)}",
    ).font = Font(size=10, color="FF5C5952")
    r += 2

    r = _title(ws, r, "EDITABLE INPUTS — every formula in this workbook reads these")
    r = _note(ws, r, "Change a value and the Cash Flow, Sources & Uses and Summary tabs recalculate.")
    r += 1

    editable = [
        ("cost", ["soft_cost_pct", "contingency_pct", "cost_escalation_annual", "demo_cost_psf"]),
        ("revenue", ["rent_psf_residential_monthly", "stabilized_occupancy", "opex_ratio",
                     "rent_growth_annual"]),
        ("debt", ["construction_ltc", "construction_annual_rate", "perm_ltv",
                  "perm_annual_rate", "perm_amortization_years", "perm_min_dscr"]),
        ("exit", ["exit_cap_rate", "selling_cost_pct", "target_developer_margin",
                  "discount_rate", "irr_hurdle"]),
    ]
    formats = {"percent": PCT, "money": MONEY, "rate": MONEY_2, "months": NUM,
               "years": NUM, "number": NUM_2}

    # Rent and exit cap are MARKET-backed: `proforma` reads them off MarketData, not off
    # the assumption set, so the assumption default (rent $3.20) is not what the model ran
    # with (ward 6 quotes $3.60). Seeding the sheet from the assumption would produce a
    # workbook quietly 12.5% out. `market` already carries any per-parcel override the user
    # applied, so it is the one correct source for both.
    market_backed = {
        "rent_psf_residential_monthly": market.rent_psf_residential_monthly,
        "exit_cap_rate": market.exit_cap_rate,
    }

    for group, keys in editable:
        ws.cell(row=r, column=1, value=vocab.ASSUMPTION_GROUP_LABELS[group]).font = Font(
            bold=True, size=10, color=ACCENT
        )
        r += 1
        values = dict(getattr(assumptions, group))
        values.update({k: v for k, v in market_backed.items() if k in values})
        for key in keys:
            label, kind = vocab.ASSUMPTION_FIELDS[key]
            # Percentages are stored as the fraction the engine uses and DISPLAYED as a
            # percentage by number format. Scaling the stored value instead would mean a
            # /100 in every formula that touches it — more places to get it wrong.
            r = a.label_value(r, key, label, values[key], formats.get(kind, NUM_2), INPUT_FILL)
        r += 1

    r = _title(ws, r, "PROGRAM & MARKET — fixed for this scenario")
    r = _note(ws, r, "Resolved by the engine from zoning and the submarket. Re-export to change.")
    r += 1
    fixed = [
        ("gross_sf", "Total floor area (SF)", program.gross_sf, NUM),
        ("net_rentable_sf", "Rentable area (SF)", program.net_rentable_sf, NUM),
        ("unit_count", "Units (est.)", program.unit_count, NUM),
        ("floors", "Floors", program.floors, NUM),
        ("parking_stalls", "Parking stalls", program.parking_stalls, NUM),
        ("parking_cost", "Cost per stall",
         assumptions.cost["parking_cost_per_stall"][program.parking_type], MONEY),
        # The EFFECTIVE rate: construction-type cost with the prototype's height factor
        # already applied (§5). The workbook must price the shell exactly as the engine
        # does, and highrise carries +6.25% over the same concrete a midrise uses.
        ("hard_cost_psf", f"Hard cost $/SF ({vocab.CONSTRUCTION_LABELS.get(ctype, ctype)})",
         hard_cost_psf(market.hard_cost_psf[program.construction_type], proto), MONEY_2),
        ("rent_premium", "Product rent premium", RENT_PREMIUM_FACTOR.get(proto, 1.0), MULT),
        ("cap_adjustment", "Product cap adjustment", EXIT_CAP_ADJUSTMENT.get(proto, 0.0), PCT),
        ("min_cap", "Minimum exit cap", MIN_EXIT_CAP_RATE, PCT),
        ("lot_area_sf", "Lot area (SF)", parcel.lot_area_sf, NUM),
        ("existing_sf", "Existing building (SF)", parcel.existing_building_sf, NUM),
        ("demo_on", "Demolition included (1 = yes)",
         1 if assumptions.cost.get("include_demolition") else 0, NUM),
        ("assessed", "Assessed land value", parcel.land_value or 0.0, MONEY),
    ]
    for key, label, value, fmt in fixed:
        r = a.label_value(r, key, label, value, fmt, FIXED_FILL)
    r += 1

    r = _title(ws, r, "TIMELINE — structural, re-export to change")
    r = _note(ws, r, "These set the number of month columns, so they cannot be changed in place.")
    r += 1
    for key, label, value in [
        ("predev", "Predevelopment (months)", predev),
        ("construction", "Construction (months)", construction),
        ("leaseup", "Lease-up (months)", leaseup),
        ("hold", "Hold after stabilization (months)", sale - stab),
    ]:
        r = a.label_value(r, key, label, value, NUM, FIXED_FILL)
    r += 1

    # --- derived market values, live off the editable base inputs ----------
    r = _title(ws, r, "DERIVED")
    r += 1
    r = a.label_value(
        r, "eff_rent", "Effective rent $/SF/mo",
        f"={a.at['rent_psf_residential_monthly']}*{a.at['rent_premium']}", MONEY_2,
    )
    r = a.label_value(
        r, "eff_cap", "Effective exit cap",
        f"=MAX({a.at['exit_cap_rate']}+{a.at['cap_adjustment']},{a.at['min_cap']})", PCT,
    )
    r = a.label_value(
        r, "shell", "Shell + parking cost",
        f"={a.at['gross_sf']}*{a.at['hard_cost_psf']}"
        f"+{a.at['parking_stalls']}*{a.at['parking_cost']}", MONEY,
    )
    r = a.label_value(
        r, "demo_cost", "Demolition cost",
        f"={a.at['demo_on']}*{a.at['existing_sf']}*{a.at['demo_cost_psf']}", MONEY,
    )
    r += 1

    # --- the seeded land value --------------------------------------------
    r = _title(ws, r, "RESIDUAL LAND VALUE — solved, not a formula")
    r = _note(
        ws, r,
        "The engine solves this: the land price at which levered IRR equals the return "
        "hurdle. That is circular in a spreadsheet, so it is seeded here.",
    )
    r = _note(
        ws, r,
        "TO RE-SOLVE after editing: Data > What-If Analysis > Goal Seek. Set the IRR cell "
        "on Summary to the hurdle by changing this cell.",
        warn=True,
    )
    r += 1
    r = a.label_value(r, "land", "Residual land value", land_solved, MONEY, WARN_FILL)

    # =======================================================================
    # Cash Flow
    # =======================================================================
    cws = wb.create_sheet("Cash Flow")
    c = Sheet(cws, "Cash Flow")
    cws.column_dimensions["A"].width = 30
    cws.freeze_panes = "B4"

    def col(t: int) -> str:
        return get_column_letter(2 + t)

    cws.cell(row=1, column=1, value="Monthly levered cash flow").font = Font(bold=True, size=12)
    cws.cell(row=2, column=1, value="Month").font = Font(bold=True, size=9)
    cws.cell(row=3, column=1, value="Phase").font = Font(bold=True, size=9)
    for t in range(months):
        cws.cell(row=2, column=2 + t, value=t).font = Font(size=9, color="FF8A8781")
        phase = (
            "Predev" if t < predev
            else "Construction" if t < predev + construction
            else "Lease-up" if t < stab
            else "Hold" if t < sale
            else "Sale"
        )
        cws.cell(row=3, column=2 + t, value=phase).font = Font(size=8, color="FF8A8781")

    rows: dict[str, int] = {}
    row_at = 5

    def series_row(key: str, label: str, fmt: str = MONEY, bold: bool = False) -> int:
        nonlocal row_at
        cell = cws.cell(row=row_at, column=1, value=label)
        cell.font = Font(size=10, bold=bold, color=INK)
        for t in range(months):
            cws.cell(row=row_at, column=2 + t).number_format = fmt
        rows[key] = row_at
        row_at += 1
        return rows[key]

    # Hard-cost timing helpers. The engine's S-curve is the symmetric-triangular CDF
    # (§6.2): cdf(x) = 2x² below the midpoint, 1 − 2(1−x)² above. Polynomial, so it lands in
    # the restricted vocabulary with no EXP and no lookup table.
    series_row("x0", "  s-curve x (start)", NUM_2)
    series_row("x1", "  s-curve x (end)", NUM_2)
    series_row("cdf0", "  s-curve cdf (start)", NUM_2)
    series_row("cdf1", "  s-curve cdf (end)", NUM_2)
    series_row("frac", "  s-curve draw fraction", NUM_2)
    series_row("esc", "  escalation factor", NUM_2)

    series_row("land", "Land")
    series_row("hard", "Hard cost")
    series_row("soft", "Soft cost")
    series_row("cont", "Contingency")
    series_row("cost", "Total cost", MONEY, bold=True)
    series_row("noi", "NOI")
    series_row("eq_draw", "Equity drawn")
    series_row("cum_eq", "  cumulative equity")
    series_row("draw", "Construction loan draw")
    series_row("interest", "Capitalized interest")
    series_row("balance", "Construction loan balance")
    series_row("perm_bal", "Permanent loan balance")
    series_row("pds", "Permanent debt service")
    series_row("eq_cf", "Equity cash flow", MONEY, bold=True)
    series_row("cum_cost", "Cumulative cost")
    series_row("cum_eq_cf", "Cumulative equity drawn")
    series_row("cum_dist", "Cumulative distributions")

    def ref(key: str, t: int) -> str:
        """Unqualified address — only safe INSIDE the Cash Flow sheet."""
        return f"{col(t)}{rows[key]}"

    def xref(key: str, t: int) -> str:
        """Sheet-qualified address, for formulas living on any other tab.

        Summary and Sources & Uses both reference the grid, and a bare `AX16` there
        silently resolves to that sheet's own empty AX16 rather than erroring — which is
        exactly the kind of quiet wrong number this export exists to avoid.
        """
        return f"'Cash Flow'!{col(t)}{rows[key]}"

    A = a.at  # assumption addresses

    # scalars used across the grid
    total_cost = f"SUM({col(0)}{rows['cost']}:{col(months - 1)}{rows['cost']})"
    equity_needed = f"(1-{A['construction_ltc']})*{total_cost}"
    hard_total = f"SUM({col(0)}{rows['hard']}:{col(months - 1)}{rows['hard']})"

    for t in range(months):
        in_constr = predev <= t < predev + construction
        # --- s-curve -------------------------------------------------------
        if in_constr:
            i = t - predev
            cws[ref("x0", t)] = f"={i}/{construction}"
            cws[ref("x1", t)] = f"={i + 1}/{construction}"
            for src, dst in (("x0", "cdf0"), ("x1", "cdf1")):
                x = ref(src, t)
                cws[ref(dst, t)] = f"=IF({x}<=0.5,2*{x}*{x},1-2*(1-{x})*(1-{x}))"
            cws[ref("frac", t)] = f"={ref('cdf1', t)}-{ref('cdf0', t)}"
        else:
            for key in ("x0", "x1", "cdf0", "cdf1", "frac"):
                cws[ref(key, t)] = 0
        cws[ref("esc", t)] = f"=(1+{A['cost_escalation_annual']})^({t}/12)"

        # --- costs (§6.2) ---------------------------------------------------
        cws[ref("land", t)] = f"={A['land']}" if t == 0 else 0
        if in_constr:
            demo = f"+{A['demo_cost']}*{ref('esc', t)}" if t == predev else ""
            cws[ref("hard", t)] = (
                f"={A['shell']}*{ref('frac', t)}*{ref('esc', t)}{demo}"
            )
        else:
            cws[ref("hard", t)] = 0
        # Soft cost spreads evenly across predevelopment + construction; contingency
        # follows the hard-cost curve exactly.
        cws[ref("soft", t)] = (
            f"={A['soft_cost_pct']}*{hard_total}/{predev + construction}"
            if t < predev + construction else 0
        )
        cws[ref("cont", t)] = (
            f"={A['contingency_pct']}*{ref('hard', t)}" if in_constr else 0
        )
        cws[ref("cost", t)] = (
            f"={ref('land', t)}+{ref('hard', t)}+{ref('soft', t)}+{ref('cont', t)}"
        )

        # --- revenue (§6.3) -------------------------------------------------
        if t >= predev + construction:
            elapsed = t - (predev + construction) + 1
            leased = (
                "1"
                if t >= stab
                else f"MIN({elapsed}/{leaseup},1)"
            )
            cws[ref("noi", t)] = (
                f"={A['net_rentable_sf']}*{A['eff_rent']}*12/12"
                f"*(1+{A['rent_growth_annual']})^({t}/12)"
                f"*{leased}*{A['stabilized_occupancy']}*(1-{A['opex_ratio']})"
            )
        else:
            cws[ref("noi", t)] = 0

        # --- construction loan (§6.4), equity first, interest capitalized ----
        if t < stab:
            prev_cum = f"{ref('cum_eq', t - 1)}" if t else "0"
            prev_bal = f"{ref('balance', t - 1)}" if t else "0"
            cws[ref("eq_draw", t)] = (
                f"=MIN({ref('cost', t)},MAX({equity_needed}-{prev_cum},0))"
            )
            cws[ref("cum_eq", t)] = f"={prev_cum}+{ref('eq_draw', t)}"
            cws[ref("draw", t)] = f"={ref('cost', t)}-{ref('eq_draw', t)}"
            cws[ref("interest", t)] = f"={prev_bal}*{A['construction_annual_rate']}/12"
            cws[ref("balance", t)] = (
                f"={prev_bal}+{ref('draw', t)}+{ref('interest', t)}"
            )
        else:
            for key in ("eq_draw", "draw", "interest"):
                cws[ref(key, t)] = 0
            cws[ref("cum_eq", t)] = f"={ref('cum_eq', t - 1)}"
            cws[ref("balance", t)] = 0

    # --- permanent takeout (§6.5) ------------------------------------------
    # Written after the grid because it needs the stabilized NOI cell to exist.
    stab_noi = f"{ref('noi', stab)}*12"
    stabilized_value = f"{stab_noi}/{A['eff_cap']}"
    p_rate = f"({A['perm_annual_rate']}/12)"
    p_n = f"({A['perm_amortization_years']}*12)"
    ltv_amount = f"{A['perm_ltv']}*{stabilized_value}"
    dscr_payment = f"{stab_noi}/{A['perm_min_dscr']}/12"
    dscr_amount = f"{dscr_payment}*(1-(1+{p_rate})^-{p_n})/{p_rate}"

    perm_row = row_at + 1
    cws.cell(row=perm_row, column=1, value="Permanent loan (at stabilization)").font = Font(
        bold=True, size=10
    )
    perm_loan_cell = f"'Cash Flow'!$B${perm_row}"
    cws.cell(row=perm_row, column=2).value = f"=MAX(MIN({ltv_amount},{dscr_amount}),0)"
    cws.cell(row=perm_row, column=2).number_format = MONEY

    pay_row = perm_row + 1
    cws.cell(row=pay_row, column=1, value="Permanent monthly payment").font = Font(size=10)
    perm_pay_cell = f"'Cash Flow'!$B${pay_row}"
    cws.cell(row=pay_row, column=2).value = (
        f"={perm_loan_cell}*{p_rate}/(1-(1+{p_rate})^-{p_n})"
    )
    cws.cell(row=pay_row, column=2).number_format = MONEY

    for t in range(months):
        if t < stab:
            cws[ref("perm_bal", t)] = 0
            cws[ref("pds", t)] = 0
        elif t == stab:
            cws[ref("perm_bal", t)] = f"={perm_loan_cell}"
            cws[ref("pds", t)] = 0
        else:
            prev = ref("perm_bal", t - 1)
            cws[ref("pds", t)] = f"={perm_pay_cell}"
            cws[ref("perm_bal", t)] = (
                f"=MAX({prev}+{prev}*{p_rate}-{perm_pay_cell},0)"
            )

    # --- equity cash flow (§6.6/6.7) ---------------------------------------
    net_sale = (
        f"({ref('noi', sale)}*12/{A['eff_cap']})*(1-{A['selling_cost_pct']})"
        f"-{ref('perm_bal', sale)}"
    )
    for t in range(months):
        parts = []
        if t < stab:
            parts.append(f"-{ref('eq_draw', t)}+{ref('noi', t)}")
        elif t == stab:
            # The construction balance is retired by the takeout; a shortfall is cash-in.
            parts.append(
                f"{perm_loan_cell}-{ref('balance', stab - 1)}+{ref('noi', t)}"
            )
        else:
            parts.append(f"{ref('noi', t)}-{ref('pds', t)}")
        if t == sale:
            parts.append(f"+({net_sale})")
        cws[ref("eq_cf", t)] = "=" + "".join(
            p if p.startswith(("+", "-")) else p for p in parts
        )

        cws[ref("cum_cost", t)] = (
            f"={ref('cost', t)}" if t == 0 else f"={ref('cum_cost', t - 1)}+{ref('cost', t)}"
        )
        # Contributions and distributions are accumulated separately because the engine's
        # equity multiple is distributions / contributions, not a net figure.
        eq_out = f"MAX(-{ref('eq_cf', t)},0)"
        eq_in = f"MAX({ref('eq_cf', t)},0)"
        cws[ref("cum_eq_cf", t)] = (
            f"={eq_out}" if t == 0 else f"={ref('cum_eq_cf', t - 1)}+{eq_out}"
        )
        cws[ref("cum_dist", t)] = (
            f"={eq_in}" if t == 0 else f"={ref('cum_dist', t - 1)}+{eq_in}"
        )

    # =======================================================================
    # Sources & Uses
    # =======================================================================
    sws = wb.create_sheet("Sources & Uses")
    s = Sheet(sws, "Sources & Uses")
    sws.column_dimensions["A"].width = 34
    sws.column_dimensions["B"].width = 18
    sws.column_dimensions["C"].width = 60

    def total_of(key: str) -> str:
        return f"SUM('Cash Flow'!{col(0)}{rows[key]}:{col(months - 1)}{rows[key]})"

    r = 1
    r = _title(sws, r, "USES")
    r += 1
    r = s.label_value(r, "u_hard", "Construction", f"={total_of('hard')}", MONEY)
    r = s.label_value(r, "u_soft", "Soft costs", f"={total_of('soft')}", MONEY)
    r = s.label_value(r, "u_cont", "Contingency", f"={total_of('cont')}", MONEY)
    r = s.label_value(r, "u_land", "Land", f"={total_of('land')}", MONEY)
    r = s.label_value(r, "u_int", "Loan interest", f"={total_of('interest')}", MONEY)
    r = s.label_value(
        r, "uses_total", "Total uses",
        f"={s.at['u_hard']}+{s.at['u_soft']}+{s.at['u_cont']}+{s.at['u_land']}+{s.at['u_int']}",
        MONEY,
    )
    r += 1

    r = _title(sws, r, "SOURCES")
    r += 1
    # Reproduces `serializers.underwrite_response` exactly: the loan is principal draws PLUS
    # capitalized interest, and equity is the cash put in BEFORE stabilization.
    #
    # The development cut-off is the point (§6.4). A cash-in refinancing at takeout and any
    # operating shortfall during hold are real capital but are not sources that funded the
    # build, and they have no matching entry in uses — counting them is what used to make
    # the two sides disagree.
    r = s.label_value(
        r, "s_loan", "Construction loan",
        f"={total_of('draw')}+{total_of('interest')}", MONEY,
    )
    r = s.label_value(
        r, "s_equity", "Equity", f"={xref('cum_eq_cf', stab - 1)}", MONEY,
    )
    r = s.label_value(
        r, "sources_total", "Total sources", f"={s.at['s_loan']}+{s.at['s_equity']}", MONEY,
    )
    r += 1

    r = s.label_value(
        r, "balance_check", "Sources − uses",
        f"={s.at['sources_total']}-{s.at['uses_total']}", MONEY, WARN_FILL,
    )
    r = _note(
        sws, r,
        "Sources and uses are the same money counted twice, so this reads $0 — including "
        "after you edit an input above. If it ever does not, the capital stack has not "
        "closed and the figures beside it should not be relied on.",
    )

    # =======================================================================
    # Summary
    # =======================================================================
    mws = wb.create_sheet("Summary")
    m = Sheet(mws, "Summary")
    mws.column_dimensions["A"].width = 34
    mws.column_dimensions["B"].width = 18
    mws.column_dimensions["C"].width = 22
    mws.column_dimensions["D"].width = 16

    eq_range = f"'Cash Flow'!{col(0)}{rows['eq_cf']}:{col(months - 1)}{rows['eq_cf']}"

    r = 1
    mws.cell(row=r, column=1, value="Summary").font = Font(bold=True, size=14)
    r += 1
    mws.cell(
        row=r, column=1,
        value=f"{record.get('address') or parcel.ssl} · "
              f"{vocab.PROTOTYPE_LABELS.get(proto, proto)} · "
              f"{vocab.CONSTRUCTION_LABELS.get(ctype, ctype)}",
    ).font = Font(size=10, color="FF5C5952")
    r += 2

    r = _title(mws, r, "HEADLINE")
    r += 1
    r = m.label_value(r, "rlv", "Financial feasibility (RLV)", f"={A['land']}", MONEY,
                      note="Solved by the engine — see Assumptions")
    # IRR: the engine annualizes a monthly rate, so the sheet does the same. The guess is
    # seeded from the engine's own answer so Excel converges on the same root.
    monthly_guess = (1 + (outputs.irr or 0.17)) ** (1 / 12) - 1
    r = m.label_value(
        r, "irr", "Annual return (levered IRR)",
        f"=(1+IRR({eq_range},{monthly_guess:.10f}))^12-1", PCT,
        note="Equals the hurdle when the land value above is correctly solved",
    )
    r = m.label_value(r, "hurdle", "Return hurdle", f"={A['irr_hurdle']}", PCT)
    r = m.label_value(
        r, "tdc", "Total cost to build",
        f"='Sources & Uses'!$B${_row_of(s.at['uses_total'])}", MONEY,
    )
    r = m.label_value(
        r, "noi", "Yearly income (NOI)", f"={xref('noi', stab)}*12", MONEY,
    )
    r = m.label_value(
        r, "exit_value", "Sale value at exit",
        f"={xref('noi', sale)}*12/{A['eff_cap']}", MONEY,
    )
    r = m.label_value(
        r, "equity_multiple", "Cash back on equity",
        f"={xref('cum_dist', months - 1)}/{xref('cum_eq_cf', months - 1)}", MULT,
    )
    r = m.label_value(
        r, "profit_margin", "Profit margin",
        f"=({m.at['exit_value']}-{m.at['tdc']})/{m.at['tdc']}", PCT,
    )
    r += 1

    # --- reconciliation ----------------------------------------------------
    r = _title(mws, r, "RECONCILIATION — this workbook against the engine")
    r = _note(
        mws, r,
        "Every delta is $0 on open. That is the proof the formulas reproduce the model. "
        "Change an input and the deltas move — that is the workbook working, not breaking.",
    )
    r += 1
    mws.cell(row=r, column=1, value="Metric").font = Font(bold=True, size=9)
    mws.cell(row=r, column=2, value="This workbook").font = Font(bold=True, size=9)
    mws.cell(row=r, column=3, value="Engine at export").font = Font(bold=True, size=9)
    mws.cell(row=r, column=4, value="Delta").font = Font(bold=True, size=9)
    r += 1

    recon = [
        ("Financial feasibility (RLV)", m.at["rlv"], land_solved, MONEY),
        ("Total cost to build", m.at["tdc"], outputs.total_development_cost, MONEY),
        ("Yearly income (NOI)", m.at["noi"], outputs.noi, MONEY),
        ("Sale value at exit", m.at["exit_value"], outputs.exit_value, MONEY),
        ("Annual return (levered IRR)", m.at["irr"], outputs.irr, PCT),
        ("Profit margin", m.at["profit_margin"], outputs.profit_margin, PCT),
    ]
    for label, live, engine_value, fmt in recon:
        mws.cell(row=r, column=1, value=label).font = Font(size=10)
        cell = mws.cell(row=r, column=2, value=f"={live}")
        cell.number_format = fmt
        engine_cell = mws.cell(row=r, column=3, value=engine_value)
        engine_cell.number_format = fmt
        delta = mws.cell(
            row=r, column=4,
            value=f"=$B${r}-$C${r}",
        )
        delta.number_format = fmt
        m.at[f"delta_{label}"] = f"'Summary'!$D${r}"
        r += 1

    return wb


def _row_of(address: str) -> int:
    """`'Sheet'!$B$12` -> 12. Used where a formula needs the row rather than the address."""
    return int(address.rsplit("$", 1)[1])


def filename_for(result: dict) -> str:
    """`residual-1301-delaware-ave-sw.xlsx` — named from the address, never the parcel id."""
    record = result["record"]
    raw = (record.get("address") or result["parcel"].ssl).strip().lower()
    slug = "".join(ch if ch.isalnum() else "-" for ch in raw).strip("-")[:48] or "parcel"
    return f"residual-{slug}.xlsx"
