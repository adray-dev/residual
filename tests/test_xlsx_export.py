"""The Excel export's consistency gate.

A live-formula workbook has exactly one way to betray the user: compute a different number
than the tool it came from. So this recomputes the generated file with an independent
formula engine and asserts the Summary tab equals the Python engine.

`formulas` is a Python implementation of Excel's formula language. It proves the ARITHMETIC
is right; it cannot prove Microsoft Excel agrees. That gap is managed by keeping the
workbook to `+ - * / ^`, SUM, MIN, MAX, IF and IRR — no volatile functions, no array
formulas, nothing whose edge cases differ between implementations.

Tolerances: money to the cent, rates to 1e-9. Bit-exactness is not achievable — the engine
sums numpy vectors and the sheet sums cell by cell, so association order differs in the last
few ULPs.
"""
from __future__ import annotations

import io
import os

import pytest
from fastapi.testclient import TestClient

pytestmark = pytest.mark.skipif(
    not os.environ.get("DATABASE_URL"), reason="DATABASE_URL not set"
)

formulas = pytest.importorskip("formulas", reason="the xlsx consistency check needs `formulas`")

CENT = 0.01
RATE = 1e-9


@pytest.fixture(scope="module")
def client():
    from api.main import app
    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="module")
def parcels(client):
    """A few real scored parcels that carry an assessed value."""
    rows = client.get(
        "/map/query", params={"statuses": ["scored"], "limit": 200}
    ).json()["rows"]
    return [r["parcel_id"] for r in rows if r["land_value"]][:3]


def _computed(xlsx_bytes: bytes, tmp_path) -> dict:
    """Recalculate the workbook and return {cell address: value} for the Summary tab."""
    path = tmp_path / "book.xlsx"
    path.write_bytes(xlsx_bytes)
    model = formulas.ExcelModel().loads(str(path)).finish()
    solution = model.calculate()

    out: dict[str, float] = {}
    for key, value in solution.items():
        upper = key.upper()
        if "SUMMARY'!" not in upper:
            continue
        cell = upper.rsplit("!", 1)[1]
        try:
            out[cell] = float(getattr(value, "value", value).ravel()[0])
        except Exception:                      # noqa: BLE001 — non-numeric cells are labels
            continue
    return out


def _labelled(xlsx_bytes: bytes) -> dict[str, str]:
    """Map each Summary label to the address of its value cell, so the test reads by NAME.

    Pinning cell addresses in the test would make it a change-detector for layout; what is
    under test is that the number beside "Total cost to build" is right.
    """
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(xlsx_bytes))["Summary"]
    found: dict[str, str] = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        label = row[0].value
        if isinstance(label, str) and label.strip():
            found.setdefault(label.strip(), f"B{row[0].row}")
    return found


def _export(client, parcel_id: str, **body) -> bytes:
    response = client.post("/export.xlsx", json={"parcel_id": parcel_id, **body})
    assert response.status_code == 200, response.text
    assert "spreadsheetml" in response.headers["content-type"]
    return response.content


def _engine(client, parcel_id: str, **body) -> dict:
    if body:
        return client.post(f"/parcel/{parcel_id}/underwrite", json=body).json()
    return client.get(f"/parcel/{parcel_id}/underwrite").json()


def _assert_matches(xlsx: bytes, engine: dict, tmp_path):
    cells = _computed(xlsx, tmp_path)
    where = _labelled(xlsx)

    def sheet(label: str) -> float:
        address = where[label]
        assert address in cells, f"{label} at {address} did not compute"
        return cells[address]

    returns = engine["returns"]
    assert sheet("Financial feasibility (RLV)") == pytest.approx(
        engine["feasibility_value"]["full"], abs=CENT
    )
    assert sheet("Total cost to build") == pytest.approx(
        returns["total_development_cost"], abs=CENT
    )
    assert sheet("Yearly income (NOI)") == pytest.approx(returns["noi"], abs=CENT)
    assert sheet("Sale value at exit") == pytest.approx(returns["exit_value"], abs=CENT)
    assert sheet("Income vs cost (yield on cost)") == pytest.approx(
        returns["yield_on_cost"], abs=RATE
    )
    assert sheet("Profit margin") == pytest.approx(returns["profit_margin"], abs=RATE)
    # At the solved land value the IRR is the hurdle by construction — that identity is the
    # in-sheet proof that the seeded RLV is the right one.
    assert sheet("Annual return (levered IRR)") == pytest.approx(
        returns["target_return"], abs=1e-6
    )


def test_default_scenario_workbook_matches_the_engine(client, parcels, tmp_path):
    for parcel_id in parcels:
        _assert_matches(_export(client, parcel_id), _engine(client, parcel_id), tmp_path)


def test_edited_scenario_workbook_matches_the_engine(client, parcels, tmp_path):
    """An edit has to flow through every formula, not just the input cell."""
    body = {"exit": {"exit_cap_rate": 0.05}, "debt": {"construction_ltc": 0.70}}
    parcel_id = parcels[0]
    _assert_matches(_export(client, parcel_id, **body), _engine(client, parcel_id, **body), tmp_path)


def test_demolition_scenario_workbook_matches_the_engine(client, parcels, tmp_path):
    """Demolition adds a cost in the first construction month only."""
    body = {"include_demolition": True}
    parcel_id = parcels[0]
    _assert_matches(_export(client, parcel_id, **body), _engine(client, parcel_id, **body), tmp_path)


def test_sources_and_uses_reproduces_the_engine_including_its_imbalance(
    client, parcels, tmp_path
):
    """The workbook must agree with the TOOL, not with an idealised model.

    Sources and uses currently diverge on edited scenarios (KNOWN_ISSUES). The export
    reproduces that rather than quietly computing something the app does not show, and the
    gap is surfaced in a check cell.
    """
    body = {"debt": {"construction_ltc": 0.70}}
    parcel_id = parcels[0]
    xlsx = _export(client, parcel_id, **body)
    engine = _engine(client, parcel_id, **body)

    path = tmp_path / "su.xlsx"
    path.write_bytes(xlsx)
    solution = formulas.ExcelModel().loads(str(path)).finish().calculate()

    values: dict[str, float] = {}
    for key, value in solution.items():
        upper = key.upper()
        if "SOURCES & USES'!" not in upper:
            continue
        try:
            values[upper.rsplit("!", 1)[1]] = float(getattr(value, "value", value).ravel()[0])
        except Exception:                      # noqa: BLE001
            continue

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(xlsx))["Sources & Uses"]
    where = {
        row[0].value.strip(): f"B{row[0].row}"
        for row in ws.iter_rows(min_col=1, max_col=1)
        if isinstance(row[0].value, str) and row[0].value.strip()
    }

    su = engine["sources_uses"]
    assert values[where["Total uses"]] == pytest.approx(su["uses_total"], abs=CENT)
    assert values[where["Total sources"]] == pytest.approx(su["sources_total"], abs=CENT)
    assert values[where["Sources − uses"]] == pytest.approx(
        su["sources_total"] - su["uses_total"], abs=CENT
    )


def test_export_needs_no_saved_scenario(client, parcels):
    """Saving is for keeping a scenario, not a precondition for getting a workbook."""
    before = len(client.get("/scenarios").json())
    _export(client, parcels[0])
    assert len(client.get("/scenarios").json()) == before


def test_export_is_named_after_the_address_not_the_parcel_id(client, parcels):
    response = client.post("/export.xlsx", json={"parcel_id": parcels[0]})
    disposition = response.headers["content-disposition"]
    assert disposition.startswith("attachment;") and disposition.endswith('.xlsx"')
    assert "ssl" not in disposition.lower()


def test_an_unmodellable_parcel_is_refused_rather_than_exported(client):
    rows = client.get("/map/query", params={"statuses": ["exempt"], "limit": 1}).json()["rows"]
    response = client.post("/export.xlsx", json={"parcel_id": rows[0]["parcel_id"]})
    assert response.status_code == 422
